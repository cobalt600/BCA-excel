# -*- coding: utf-8 -*-

#プレートリーダーから吐き出されたエクセルファイルから必要な情報を抜き出すプログラム
#吸光度とかはいらないけど、最小二乗法でフィッティングされた濃度（平均が取れればなおよしだが、場所がコロコロ変わる）とr，r^2が最低限必要
#余力があればグラフも吐き出したい
#最終的には【印刷して】実験ノートに貼る必要がある（左右の余白を確保してA4に印刷、上半分に全部の情報が収まっていると使い勝手が良い）
#最終アウトプットはxlsxでもwordでもPDFでもいい。改変可能性を減らすためにはPDFが良いかな？
#同時並行でGitの使い方を理解したい、ので、とりあえずネットにあった適当なコードを貼り付けたものを初版として、これを改変していく。

#絶妙のpandasだとやりたいことと乖離する気がしたので、書き直す

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment

import xlwings as xw
import zipfile
import shutil
import os

#当該のxlsxをZIP形式のファイルとして開いてグラフ（image.png）を保存する
xlsx_zip = zipfile.ZipFile("sample.xlsx")
xlsx_zip.extract("xl/media/image1.png")
xlsx_zip.close()

#sample.xlsxを開いて、Result（結果をまとめるシート）を削除して作成
filepath = 'sample.xlsx'
wb = openpyxl.load_workbook(filepath)
wb.remove(wb["Result"])
ws_new = wb.create_sheet(index=0, title='Result')

#とりあえず必要なシートを定義、ws1に情報を集める
ws1 = wb["Result"]
ws2 = wb["Linear regression fit"]
ws3 = wb["End point_1"]

#Linear~のA3:A10のRun infoをコピー、ResultシートのA1からに貼り付ける
#この書き方は良くない気がしている（行ごと取得するのがスマートなのであとで直す）
for i in range(3,10):
    copy = ws2.cell(row = i, column = 1).value
    ws1.cell(row = i-2, column = 1, value = copy)
i += 1

#Linear~のA12:C18のCurve infoをコピー、ResultシートのF1からに貼り付ける
#この書き方は良くない気がしている（行ごと取得するのがスマートなのであとで直す）
for i in range(15,19):
    for j in range(2,4):
        copy = ws2.cell(row = i, column = j).value
        ws1.cell(row = i-11, column = j+4, value = copy)
i += 1
j += 1

#End point_1のA48:M56のConc dataをコピー、ResultシートのA9からに貼り付ける
#この書き方は良くない気がしている（行ごと取得するのがスマートなのであとで直す）
for i in range(48,57):
    for j in range(1,14):
        copy = ws3.cell(row = i, column = j).value
        ws1.cell(row = i-39, column = j, value = copy)
i += 1
j += 1

#罫線を追加
sheet = wb['Result']
side1 = Side(style='thin', color='000000')
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)

for row in sheet['B10:M17']:
    for cell in row:
        cell.border = border_aro

#プレートの番地を太字でセンタリングにする
sheet = wb['Result']
for row in ws1["A9:M9"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="centerContinuous")
        cell.font = Font(bold=True)

for row in ws1["A10:A17"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="centerContinuous")
        cell.font = Font(bold=True)

#最初に保存したグラフ（画像）を指定してリサイズして貼り付け
img = Image('xl/media/image1.png')
img.width = 330
img.height = 110
ws1.add_image(img, 'H2')

#行の高さと幅の調整
sheet.column_dimensions['A'].width = 3.5
sheet.column_dimensions['B'].width = 7
sheet.column_dimensions['C'].width = 7
sheet.column_dimensions['D'].width = 7
sheet.column_dimensions['E'].width = 7
sheet.column_dimensions['F'].width = 7
sheet.column_dimensions['G'].width = 7
sheet.column_dimensions['H'].width = 7
sheet.column_dimensions['I'].width = 7
sheet.column_dimensions['J'].width = 7
sheet.column_dimensions['K'].width = 7
sheet.column_dimensions['L'].width = 7
sheet.column_dimensions['M'].width = 7

#印刷領域の指定（列が1枚に収まるように）して上書き保存
ws1.page_setup.fitToWidth = 1
ws1.page_setup.fitToHeight = 0
ws1.sheet_properties.pageSetUpPr.fitToPage = True
wb.save(filepath)

#xlwingsを使ってPDF出力
wb = xw.Book('sample.xlsx')
sheet = wb.sheets['Result']
path = os.getcwd()
sheet.to_pdf('sample.pdf')
wb.close()

#最初に作られたImage.pngが保存されているディレクトリを削除する
shutil.rmtree('xl/')
